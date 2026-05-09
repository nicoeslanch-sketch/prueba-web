"""
Sistema Financiero PyME - Generador de Excel
Script para crear una plantilla Excel profesional para PyMEs con estados financieros y dashboard.
Usa openpyxl para generar el archivo .xlsx.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule

# Crear workbook
wb = Workbook()

# Crear hojas
sheets = ['Portada', 'Instrucciones', 'Parametros', 'Carga_Mensual', 'Estado_Resultados', 'Balance_General', 'Flujo_Efectivo', 'Cambios_Patrimonio', 'Ratios_KPI', 'Dashboard']
for sheet in sheets:
    wb.create_sheet(sheet)

# Remover hoja por defecto
if 'Sheet' in wb.sheetnames:
    wb.remove(wb['Sheet'])

# Estilos comunes
header_font = Font(bold=True, size=12, color="FFFFFF")
header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
body_font = Font(size=10)
currency_style = NamedStyle(name="currency", number_format='"CLP" #,##0')
percentage_style = NamedStyle(name="percentage", number_format="0.00%")

wb.add_named_style(currency_style)
wb.add_named_style(percentage_style)

# Función para aplicar bordes
def apply_border(ws, range_str):
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for row in ws[range_str]:
        for cell in row:
            cell.border = thin_border

# Función para congelar encabezados
def freeze_panes(ws, row, col):
    ws.freeze_panes = ws.cell(row=row, column=col)

# Hoja 1: Portada
ws_portada = wb['Portada']
ws_portada['A1'] = "Sistema Financiero PyME"
ws_portada['A1'].font = Font(bold=True, size=24, color="1F497D")
ws_portada['A1'].alignment = Alignment(horizontal="center")
ws_portada.merge_cells('A1:G1')

ws_portada['A3'] = "Estado de Resultados, Balance General, Flujo de Efectivo, Estado de Cambios en el Patrimonio y Dashboard Ejecutivo"
ws_portada['A3'].font = Font(size=14, color="1F497D")
ws_portada['A3'].alignment = Alignment(horizontal="center")
ws_portada.merge_cells('A3:G3')

ws_portada['A5'] = "Propuesta de Valor:"
ws_portada['A5'].font = Font(bold=True, size=12)
ws_portada['A7'] = "• Toma decisiones con datos reales, no con intuición."
ws_portada['A8'] = "• Entiende la salud financiera de tu negocio mes a mes."
ws_portada['A9'] = "• Herramienta diseñada para PyMEs."
ws_portada['A7'].font = body_font
ws_portada['A8'].font = body_font
ws_portada['A9'].font = body_font

# Fondo celeste suave
for row in range(1, 20):
    for col in range(1, 8):
        cell = ws_portada.cell(row=row, column=col)
        cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")

# Hoja 2: Instrucciones
ws_inst = wb['Instrucciones']
ws_inst['A1'] = "Instrucciones de Uso"
ws_inst['A1'].font = header_font
ws_inst['A1'].fill = header_fill
ws_inst['A1'].alignment = Alignment(horizontal="center")
ws_inst.merge_cells('A1:D1')

ws_inst['A3'] = "1. Comienza por la hoja 'Parametros' para configurar el año, moneda y otros datos básicos."
ws_inst['A4'] = "2. En 'Carga_Mensual', ingresa los datos mensuales de tu empresa."
ws_inst['A5'] = "3. Las hojas de estados financieros se calculan automáticamente."
ws_inst['A6'] = "4. Revisa el 'Dashboard' para un resumen ejecutivo."
ws_inst['A7'] = "5. Las celdas en azul claro son editables; las demás están protegidas por fórmulas."
ws_inst['A8'] = "6. Asegúrate de que el balance cuadre antes de confiar en los números."

# Hoja 3: Parametros
ws_param = wb['Parametros']
ws_param['A1'] = "Parámetros Generales"
ws_param['A1'].font = header_font
ws_param['A1'].fill = header_fill
ws_param.merge_cells('A1:B1')

ws_param['A2'] = "Año"
ws_param['B2'] = 2024  # Editable
ws_param['A3'] = "Moneda"
ws_param['B3'] = "CLP"
ws_param['A4'] = "Tasa de Impuesto (%)"
ws_param['B4'] = 0.25  # 25%
ws_param['A5'] = "Nombre Empresa"
ws_param['B5'] = "Mi Empresa PyME"

# Aplicar estilos
for row in range(2, 6):
    ws_param.cell(row=row, column=2).fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")

# Hoja 4: Carga_Mensual
ws_carga = wb['Carga_Mensual']
meses = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio', 'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre', 'Total Anual']

# Encabezados
ws_carga['A1'] = "Concepto"
ws_carga['A1'].font = header_font
ws_carga['A1'].fill = header_fill
for i, mes in enumerate(meses, start=2):
    ws_carga.cell(row=1, column=i, value=mes).font = header_font
    ws_carga.cell(row=1, column=i).fill = header_fill

# Datos de entrada
carga_data = [
    "Ventas brutas",
    "Descuentos y devoluciones",
    "Costo de ventas",
    "Gastos de venta / comerciales",
    "Gastos administrativos",
    "Remuneraciones",
    "Arriendo",
    "Servicios básicos",
    "Marketing / publicidad",
    "Transporte / logística",
    "Depreciación y amortización",
    "Otros gastos operacionales",
    "Otros ingresos",
    "Gastos financieros / intereses",
    "Impuesto a la renta",
    "Cobros a clientes",
    "Pago a proveedores",
    "Pago de remuneraciones",
    "Pago de gastos operacionales",
    "Pago de impuestos",
    "Compra de activos fijos",
    "Venta de activos",
    "Préstamos recibidos",
    "Pago de deuda",
    "Aportes de capital",
    "Retiros / dividendos",
    "Caja inicial del mes",
    "Cuentas por cobrar",
    "Inventario",
    "Otros activos corrientes",
    "Activo fijo bruto",
    "Depreciación acumulada",
    "Otros activos no corrientes",
    "Proveedores",
    "Deuda corto plazo",
    "Impuestos por pagar",
    "Otros pasivos corrientes",
    "Deuda largo plazo",
    "Otros pasivos no corrientes",
    "Capital aportado",
    "Reservas",
    "Utilidades retenidas iniciales"
]

for row, concepto in enumerate(carga_data, start=2):
    ws_carga.cell(row=row, column=1, value=concepto).font = body_font
    for col in range(2, 15):  # Meses + total
        cell = ws_carga.cell(row=row, column=col)
        cell.fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
        if col == 14:  # Total Anual
            cell.value = f"=SUM(B{row}:M{row})"

# Congelar encabezados
freeze_panes(ws_carga, 2, 2)

# Ajustar anchos
ws_carga.column_dimensions['A'].width = 30
for col in range(2, 15):
    ws_carga.column_dimensions[get_column_letter(col)].width = 15

# Hoja 5: Estado_Resultados
ws_er = wb['Estado_Resultados']
# Encabezados
ws_er['A1'] = "Estado de Resultados"
ws_er['A1'].font = header_font
ws_er['A1'].fill = header_fill
ws_er.merge_cells('A1:N1')

for i, mes in enumerate(meses, start=2):
    ws_er.cell(row=2, column=i, value=mes).font = header_font
    ws_er.cell(row=2, column=i).fill = header_fill

# Conceptos del Estado de Resultados
er_conceptos = [
    ("Ventas brutas", ["=Carga_Mensual!B2", "=Carga_Mensual!C2", "=Carga_Mensual!D2", "=Carga_Mensual!E2", "=Carga_Mensual!F2", "=Carga_Mensual!G2", "=Carga_Mensual!H2", "=Carga_Mensual!I2", "=Carga_Mensual!J2", "=Carga_Mensual!K2", "=Carga_Mensual!L2", "=Carga_Mensual!M2", "=SUM(B3:M3)"]),
    ("(-) Descuentos y devoluciones", ["=Carga_Mensual!B3", "=Carga_Mensual!C3", "=Carga_Mensual!D3", "=Carga_Mensual!E3", "=Carga_Mensual!F3", "=Carga_Mensual!G3", "=Carga_Mensual!H3", "=Carga_Mensual!I3", "=Carga_Mensual!J3", "=Carga_Mensual!K3", "=Carga_Mensual!L3", "=Carga_Mensual!M3", "=SUM(B4:M4)"]),
    ("= Ventas netas", ["=B3-B4", "=C3-C4", "=D3-D4", "=E3-E4", "=F3-F4", "=G3-G4", "=H3-H4", "=I3-I4", "=J3-J4", "=K3-K4", "=L3-L4", "=M3-M4", "=SUM(B5:M5)"]),
    ("(-) Costo de ventas", ["=Carga_Mensual!B4", "=Carga_Mensual!C4", "=Carga_Mensual!D4", "=Carga_Mensual!E4", "=Carga_Mensual!F4", "=Carga_Mensual!G4", "=Carga_Mensual!H4", "=Carga_Mensual!I4", "=Carga_Mensual!J4", "=Carga_Mensual!K4", "=Carga_Mensual!L4", "=Carga_Mensual!M4", "=SUM(B6:M6)"]),
    ("= Utilidad bruta", ["=B5-B6", "=C5-C6", "=D5-D6", "=E5-E6", "=F5-F6", "=G5-G6", "=H5-H6", "=I5-I6", "=J5-J6", "=K5-K6", "=L5-L6", "=M5-M6", "=SUM(B7:M7)"]),
    ("(-) Gastos de venta", ["=Carga_Mensual!B5", "=Carga_Mensual!C5", "=Carga_Mensual!D5", "=Carga_Mensual!E5", "=Carga_Mensual!F5", "=Carga_Mensual!G5", "=Carga_Mensual!H5", "=Carga_Mensual!I5", "=Carga_Mensual!J5", "=Carga_Mensual!K5", "=Carga_Mensual!L5", "=Carga_Mensual!M5", "=SUM(B8:M8)"]),
    ("(-) Gastos administrativos", ["=Carga_Mensual!B6", "=Carga_Mensual!C6", "=Carga_Mensual!D6", "=Carga_Mensual!E6", "=Carga_Mensual!F6", "=Carga_Mensual!G6", "=Carga_Mensual!H6", "=Carga_Mensual!I6", "=Carga_Mensual!J6", "=Carga_Mensual!K6", "=Carga_Mensual!L6", "=Carga_Mensual!M6", "=SUM(B9:M9)"]),
    ("(-) Remuneraciones", ["=Carga_Mensual!B7", "=Carga_Mensual!C7", "=Carga_Mensual!D7", "=Carga_Mensual!E7", "=Carga_Mensual!F7", "=Carga_Mensual!G7", "=Carga_Mensual!H7", "=Carga_Mensual!I7", "=Carga_Mensual!J7", "=Carga_Mensual!K7", "=Carga_Mensual!L7", "=Carga_Mensual!M7", "=SUM(B10:M10)"]),
    ("(-) Otros gastos operacionales", ["=Carga_Mensual!B12", "=Carga_Mensual!C12", "=Carga_Mensual!D12", "=Carga_Mensual!E12", "=Carga_Mensual!F12", "=Carga_Mensual!G12", "=Carga_Mensual!H12", "=Carga_Mensual!I12", "=Carga_Mensual!J12", "=Carga_Mensual!K12", "=Carga_Mensual!L12", "=Carga_Mensual!M12", "=SUM(B11:M11)"]),
    ("= Utilidad operacional", ["=B7-B8-B9-B10-B11", "=C7-C8-C9-C10-C11", "=D7-D8-D9-D10-D11", "=E7-E8-E9-E10-E11", "=F7-F8-F9-F10-F11", "=G7-G8-G9-G10-G11", "=H7-H8-H9-H10-H11", "=I7-I8-I9-I10-I11", "=J7-J8-J9-J10-J11", "=K7-K8-K9-K10-K11", "=L7-L8-L9-L10-L11", "=M7-M8-M9-M10-M11", "=SUM(B12:M12)"]),
    ("(+) Otros ingresos", ["=Carga_Mensual!B13", "=Carga_Mensual!C13", "=Carga_Mensual!D13", "=Carga_Mensual!E13", "=Carga_Mensual!F13", "=Carga_Mensual!G13", "=Carga_Mensual!H13", "=Carga_Mensual!I13", "=Carga_Mensual!J13", "=Carga_Mensual!K13", "=Carga_Mensual!L13", "=Carga_Mensual!M13", "=SUM(B13:M13)"]),
    ("(-) Gastos financieros", ["=Carga_Mensual!B14", "=Carga_Mensual!C14", "=Carga_Mensual!D14", "=Carga_Mensual!E14", "=Carga_Mensual!F14", "=Carga_Mensual!G14", "=Carga_Mensual!H14", "=Carga_Mensual!I14", "=Carga_Mensual!J14", "=Carga_Mensual!K14", "=Carga_Mensual!L14", "=Carga_Mensual!M14", "=SUM(B14:M14)"]),
    ("= Utilidad antes de impuestos", ["=B12+B13-B14", "=C12+C13-C14", "=D12+D13-D14", "=E12+E13-E14", "=F12+F13-F14", "=G12+G13-G14", "=H12+H13-H14", "=I12+I13-I14", "=J12+J13-J14", "=K12+K13-K14", "=L12+L13-L14", "=M12+M13-M14", "=SUM(B15:M15)"]),
    ("(-) Impuesto a la renta", ["=Carga_Mensual!B15", "=Carga_Mensual!C15", "=Carga_Mensual!D15", "=Carga_Mensual!E15", "=Carga_Mensual!F15", "=Carga_Mensual!G15", "=Carga_Mensual!H15", "=Carga_Mensual!I15", "=Carga_Mensual!J15", "=Carga_Mensual!K15", "=Carga_Mensual!L15", "=Carga_Mensual!M15", "=SUM(B16:M16)"]),
    ("= Utilidad neta", ["=B15-B16", "=C15-C16", "=D15-D16", "=E15-E16", "=F15-F16", "=G15-G16", "=H15-H16", "=I15-I16", "=J15-J16", "=K15-K16", "=L15-L16", "=M15-M16", "=SUM(B17:M17)"])
]

for row, (concepto, formulas) in enumerate(er_conceptos, start=3):
    ws_er.cell(row=row, column=1, value=concepto).font = body_font
    for col, formula in enumerate(formulas, start=2):
        cell = ws_er.cell(row=row, column=col, value=formula)
        if col == 14:  # Total Anual
            cell.style = "currency"
        else:
            cell.style = "currency"

# Aplicar bordes
apply_border(ws_er, 'A1:N17')

# Congelar encabezados
freeze_panes(ws_er, 3, 2)

# Ajustar anchos
ws_er.column_dimensions['A'].width = 30
for col in range(2, 15):
    ws_er.column_dimensions[get_column_letter(col)].width = 15

# Hoja 6: Balance_General
ws_bg = wb['Balance_General']
ws_bg['A1'] = "Balance General"
ws_bg['A1'].font = header_font
ws_bg['A1'].fill = header_fill
ws_bg.merge_cells('A1:N1')

for i, mes in enumerate(meses, start=2):
    ws_bg.cell(row=2, column=i, value=mes).font = header_font
    ws_bg.cell(row=2, column=i).fill = header_fill

# Activos Corrientes
ws_bg['A3'] = "ACTIVOS CORRIENTES"
ws_bg['A3'].font = Font(bold=True, size=11, color="FFFFFF")
ws_bg['A3'].fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
ws_bg.merge_cells('A3:N3')

bg_activos_corrientes = [
    ("Caja y bancos", ["=Carga_Mensual!B27", "=Carga_Mensual!C27", "=Carga_Mensual!D27", "=Carga_Mensual!E27", "=Carga_Mensual!F27", "=Carga_Mensual!G27", "=Carga_Mensual!H27", "=Carga_Mensual!I27", "=Carga_Mensual!J27", "=Carga_Mensual!K27", "=Carga_Mensual!L27", "=Carga_Mensual!M27", "=SUM(B4:M4)"]),
    ("Cuentas por cobrar", ["=Carga_Mensual!B28", "=Carga_Mensual!C28", "=Carga_Mensual!D28", "=Carga_Mensual!E28", "=Carga_Mensual!F28", "=Carga_Mensual!G28", "=Carga_Mensual!H28", "=Carga_Mensual!I28", "=Carga_Mensual!J28", "=Carga_Mensual!K28", "=Carga_Mensual!L28", "=Carga_Mensual!M28", "=SUM(B5:M5)"]),
    ("Inventario", ["=Carga_Mensual!B29", "=Carga_Mensual!C29", "=Carga_Mensual!D29", "=Carga_Mensual!E29", "=Carga_Mensual!F29", "=Carga_Mensual!G29", "=Carga_Mensual!H29", "=Carga_Mensual!I29", "=Carga_Mensual!J29", "=Carga_Mensual!K29", "=Carga_Mensual!L29", "=Carga_Mensual!M29", "=SUM(B6:M6)"]),
    ("Otros activos corrientes", ["=Carga_Mensual!B30", "=Carga_Mensual!C30", "=Carga_Mensual!D30", "=Carga_Mensual!E30", "=Carga_Mensual!F30", "=Carga_Mensual!G30", "=Carga_Mensual!H30", "=Carga_Mensual!I30", "=Carga_Mensual!J30", "=Carga_Mensual!K30", "=Carga_Mensual!L30", "=Carga_Mensual!M30", "=SUM(B7:M7)"]),
    ("TOTAL ACTIVOS CORRIENTES", ["=SUM(B4:B7)", "=SUM(C4:C7)", "=SUM(D4:D7)", "=SUM(E4:E7)", "=SUM(F4:F7)", "=SUM(G4:G7)", "=SUM(H4:H7)", "=SUM(I4:I7)", "=SUM(J4:J7)", "=SUM(K4:K7)", "=SUM(L4:L7)", "=SUM(M4:M7)", "=SUM(B8:M8)"])
]

for row, (concepto, formulas) in enumerate(bg_activos_corrientes, start=4):
    ws_bg.cell(row=row, column=1, value=concepto).font = Font(bold=True if "TOTAL" in concepto else False, size=10)
    for col, formula in enumerate(formulas, start=2):
        cell = ws_bg.cell(row=row, column=col, value=formula)
        cell.style = "currency"

# Activos No Corrientes
ws_bg['A9'] = "ACTIVOS NO CORRIENTES"
ws_bg['A9'].font = Font(bold=True, size=11, color="FFFFFF")
ws_bg['A9'].fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
ws_bg.merge_cells('A9:N9')

bg_activos_no_corrientes = [
    ("Activo fijo bruto", ["=Carga_Mensual!B31", "=Carga_Mensual!C31", "=Carga_Mensual!D31", "=Carga_Mensual!E31", "=Carga_Mensual!F31", "=Carga_Mensual!G31", "=Carga_Mensual!H31", "=Carga_Mensual!I31", "=Carga_Mensual!J31", "=Carga_Mensual!K31", "=Carga_Mensual!L31", "=Carga_Mensual!M31", "=SUM(B10:M10)"]),
    ("(-) Depreciación acumulada", ["=Carga_Mensual!B32", "=Carga_Mensual!C32", "=Carga_Mensual!D32", "=Carga_Mensual!E32", "=Carga_Mensual!F32", "=Carga_Mensual!G32", "=Carga_Mensual!H32", "=Carga_Mensual!I32", "=Carga_Mensual!J32", "=Carga_Mensual!K32", "=Carga_Mensual!L32", "=Carga_Mensual!M32", "=SUM(B11:M11)"]),
    ("= Activo fijo neto", ["=B10-B11", "=C10-C11", "=D10-D11", "=E10-E11", "=F10-F11", "=G10-G11", "=H10-H11", "=I10-I11", "=J10-J11", "=K10-K11", "=L10-L11", "=M10-M11", "=SUM(B12:M12)"]),
    ("Otros activos no corrientes", ["=Carga_Mensual!B33", "=Carga_Mensual!C33", "=Carga_Mensual!D33", "=Carga_Mensual!E33", "=Carga_Mensual!F33", "=Carga_Mensual!G33", "=Carga_Mensual!H33", "=Carga_Mensual!I33", "=Carga_Mensual!J33", "=Carga_Mensual!K33", "=Carga_Mensual!L33", "=Carga_Mensual!M33", "=SUM(B13:M13)"]),
    ("TOTAL ACTIVOS NO CORRIENTES", ["=SUM(B12:B13)", "=SUM(C12:C13)", "=SUM(D12:D13)", "=SUM(E12:E13)", "=SUM(F12:F13)", "=SUM(G12:G13)", "=SUM(H12:H13)", "=SUM(I12:I13)", "=SUM(J12:J13)", "=SUM(K12:K13)", "=SUM(L12:L13)", "=SUM(M12:M13)", "=SUM(B14:M14)"])
]

for row, (concepto, formulas) in enumerate(bg_activos_no_corrientes, start=10):
    ws_bg.cell(row=row, column=1, value=concepto).font = Font(bold=True if "TOTAL" in concepto else False, size=10)
    for col, formula in enumerate(formulas, start=2):
        cell = ws_bg.cell(row=row, column=col, value=formula)
        cell.style = "currency"

# Total Activos
ws_bg['A15'] = "TOTAL ACTIVOS"
ws_bg['A15'].font = Font(bold=True, size=12, color="FFFFFF")
ws_bg['A15'].fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
ws_bg.merge_cells('A15:N15')

for col in range(2, 15):
    ws_bg.cell(row=15, column=col).value = f"=B8+B14"
    ws_bg.cell(row=15, column=col).style = "currency"

# Pasivos Corrientes
ws_bg['A17'] = "PASIVOS CORRIENTES"
ws_bg['A17'].font = Font(bold=True, size=11, color="FFFFFF")
ws_bg['A17'].fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
ws_bg.merge_cells('A17:N17')

bg_pasivos_corrientes = [
    ("Proveedores", ["=Carga_Mensual!B34", "=Carga_Mensual!C34", "=Carga_Mensual!D34", "=Carga_Mensual!E34", "=Carga_Mensual!F34", "=Carga_Mensual!G34", "=Carga_Mensual!H34", "=Carga_Mensual!I34", "=Carga_Mensual!J34", "=Carga_Mensual!K34", "=Carga_Mensual!L34", "=Carga_Mensual!M34", "=SUM(B18:M18)"]),
    ("Deuda corto plazo", ["=Carga_Mensual!B35", "=Carga_Mensual!C35", "=Carga_Mensual!D35", "=Carga_Mensual!E35", "=Carga_Mensual!F35", "=Carga_Mensual!G35", "=Carga_Mensual!H35", "=Carga_Mensual!I35", "=Carga_Mensual!J35", "=Carga_Mensual!K35", "=Carga_Mensual!L35", "=Carga_Mensual!M35", "=SUM(B19:M19)"]),
    ("Impuestos por pagar", ["=Carga_Mensual!B36", "=Carga_Mensual!C36", "=Carga_Mensual!D36", "=Carga_Mensual!E36", "=Carga_Mensual!F36", "=Carga_Mensual!G36", "=Carga_Mensual!H36", "=Carga_Mensual!I36", "=Carga_Mensual!J36", "=Carga_Mensual!K36", "=Carga_Mensual!L36", "=Carga_Mensual!M36", "=SUM(B20:M20)"]),
    ("Otros pasivos corrientes", ["=Carga_Mensual!B37", "=Carga_Mensual!C37", "=Carga_Mensual!D37", "=Carga_Mensual!E37", "=Carga_Mensual!F37", "=Carga_Mensual!G37", "=Carga_Mensual!H37", "=Carga_Mensual!I37", "=Carga_Mensual!J37", "=Carga_Mensual!K37", "=Carga_Mensual!L37", "=Carga_Mensual!M37", "=SUM(B21:M21)"]),
    ("TOTAL PASIVOS CORRIENTES", ["=SUM(B18:B21)", "=SUM(C18:C21)", "=SUM(D18:D21)", "=SUM(E18:E21)", "=SUM(F18:F21)", "=SUM(G18:G21)", "=SUM(H18:H21)", "=SUM(I18:I21)", "=SUM(J18:J21)", "=SUM(K18:K21)", "=SUM(L18:L21)", "=SUM(M18:M21)", "=SUM(B22:M22)"])
]

for row, (concepto, formulas) in enumerate(bg_pasivos_corrientes, start=18):
    ws_bg.cell(row=row, column=1, value=concepto).font = Font(bold=True if "TOTAL" in concepto else False, size=10)
    for col, formula in enumerate(formulas, start=2):
        cell = ws_bg.cell(row=row, column=col, value=formula)
        cell.style = "currency"

# Pasivos No Corrientes
ws_bg['A23'] = "PASIVOS NO CORRIENTES"
ws_bg['A23'].font = Font(bold=True, size=11, color="FFFFFF")
ws_bg['A23'].fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
ws_bg.merge_cells('A23:N23')

bg_pasivos_no_corrientes = [
    ("Deuda largo plazo", ["=Carga_Mensual!B38", "=Carga_Mensual!C38", "=Carga_Mensual!D38", "=Carga_Mensual!E38", "=Carga_Mensual!F38", "=Carga_Mensual!G38", "=Carga_Mensual!H38", "=Carga_Mensual!I38", "=Carga_Mensual!J38", "=Carga_Mensual!K38", "=Carga_Mensual!L38", "=Carga_Mensual!M38", "=SUM(B24:M24)"]),
    ("Otros pasivos no corrientes", ["=Carga_Mensual!B39", "=Carga_Mensual!C39", "=Carga_Mensual!D39", "=Carga_Mensual!E39", "=Carga_Mensual!F39", "=Carga_Mensual!G39", "=Carga_Mensual!H39", "=Carga_Mensual!I39", "=Carga_Mensual!J39", "=Carga_Mensual!K39", "=Carga_Mensual!L39", "=Carga_Mensual!M39", "=SUM(B25:M25)"]),
    ("TOTAL PASIVOS NO CORRIENTES", ["=SUM(B24:B25)", "=SUM(C24:C25)", "=SUM(D24:D25)", "=SUM(E24:E25)", "=SUM(F24:F25)", "=SUM(G24:G25)", "=SUM(H24:H25)", "=SUM(I24:I25)", "=SUM(J24:J25)", "=SUM(K24:K25)", "=SUM(L24:L25)", "=SUM(M24:M25)", "=SUM(B26:M26)"])
]

for row, (concepto, formulas) in enumerate(bg_pasivos_no_corrientes, start=24):
    ws_bg.cell(row=row, column=1, value=concepto).font = Font(bold=True if "TOTAL" in concepto else False, size=10)
    for col, formula in enumerate(formulas, start=2):
        cell = ws_bg.cell(row=row, column=col, value=formula)
        cell.style = "currency"

# Total Pasivos
ws_bg['A27'] = "TOTAL PASIVOS"
ws_bg['A27'].font = Font(bold=True, size=12, color="FFFFFF")
ws_bg['A27'].fill = PatternFill(start_color="843C0C", end_color="843C0C", fill_type="solid")
ws_bg.merge_cells('A27:N27')

for col in range(2, 15):
    ws_bg.cell(row=27, column=col).value = f"=B22+B26"
    ws_bg.cell(row=27, column=col).style = "currency"

# Patrimonio
ws_bg['A29'] = "PATRIMONIO"
ws_bg['A29'].font = Font(bold=True, size=11, color="FFFFFF")
ws_bg['A29'].fill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
ws_bg.merge_cells('A29:N29')

bg_patrimonio = [
    ("Capital aportado", ["=Carga_Mensual!B40", "=Carga_Mensual!C40", "=Carga_Mensual!D40", "=Carga_Mensual!E40", "=Carga_Mensual!F40", "=Carga_Mensual!G40", "=Carga_Mensual!H40", "=Carga_Mensual!I40", "=Carga_Mensual!J40", "=Carga_Mensual!K40", "=Carga_Mensual!L40", "=Carga_Mensual!M40", "=SUM(B30:M30)"]),
    ("Reservas", ["=Carga_Mensual!B41", "=Carga_Mensual!C41", "=Carga_Mensual!D41", "=Carga_Mensual!E41", "=Carga_Mensual!F41", "=Carga_Mensual!G41", "=Carga_Mensual!H41", "=Carga_Mensual!I41", "=Carga_Mensual!J41", "=Carga_Mensual!K41", "=Carga_Mensual!L41", "=Carga_Mensual!M41", "=SUM(B31:M31)"]),
    ("Utilidades retenidas iniciales", ["=Carga_Mensual!B42", "=Carga_Mensual!C42", "=Carga_Mensual!D42", "=Carga_Mensual!E42", "=Carga_Mensual!F42", "=Carga_Mensual!G42", "=Carga_Mensual!H42", "=Carga_Mensual!I42", "=Carga_Mensual!J42", "=Carga_Mensual!K42", "=Carga_Mensual!L42", "=Carga_Mensual!M42", "=SUM(B32:M32)"]),
    ("(+) Utilidad neta del período", ["=Estado_Resultados!B17", "=Estado_Resultados!C17", "=Estado_Resultados!D17", "=Estado_Resultados!E17", "=Estado_Resultados!F17", "=Estado_Resultados!G17", "=Estado_Resultados!H17", "=Estado_Resultados!I17", "=Estado_Resultados!J17", "=Estado_Resultados!K17", "=Estado_Resultados!L17", "=Estado_Resultados!M17", "=Estado_Resultados!N17"]),
    ("(-) Dividendos", ["=Carga_Mensual!B26", "=Carga_Mensual!C26", "=Carga_Mensual!D26", "=Carga_Mensual!E26", "=Carga_Mensual!F26", "=Carga_Mensual!G26", "=Carga_Mensual!H26", "=Carga_Mensual!I26", "=Carga_Mensual!J26", "=Carga_Mensual!K26", "=Carga_Mensual!L26", "=Carga_Mensual!M26", "=SUM(B35:M35)"]),
    ("TOTAL PATRIMONIO", ["=SUM(B30:B34)", "=SUM(C30:C34)", "=SUM(D30:D34)", "=SUM(E30:E34)", "=SUM(F30:F34)", "=SUM(G30:G34)", "=SUM(H30:H34)", "=SUM(I30:I34)", "=SUM(J30:J34)", "=SUM(K30:K34)", "=SUM(L30:L34)", "=SUM(M30:M34)", "=SUM(B35:M35)"])
]

for row, (concepto, formulas) in enumerate(bg_patrimonio, start=30):
    ws_bg.cell(row=row, column=1, value=concepto).font = Font(bold=True if "TOTAL" in concepto else False, size=10)
    for col, formula in enumerate(formulas, start=2):
        cell = ws_bg.cell(row=row, column=col, value=formula)
        cell.style = "currency"

# Total Pasivos + Patrimonio
ws_bg['A36'] = "TOTAL PASIVOS + PATRIMONIO"
ws_bg['A36'].font = Font(bold=True, size=12, color="FFFFFF")
ws_bg['A36'].fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
ws_bg.merge_cells('A36:N36')

for col in range(2, 15):
    ws_bg.cell(row=36, column=col).value = f"=B27+B35"
    ws_bg.cell(row=36, column=col).style = "currency"

# Aplicar bordes
apply_border(ws_bg, 'A1:N36')

# Congelar encabezados
freeze_panes(ws_bg, 3, 2)

# Ajustar anchos
ws_bg.column_dimensions['A'].width = 30
for col in range(2, 15):
    ws_bg.column_dimensions[get_column_letter(col)].width = 15

# Hoja 7: Flujo_Efectivo
ws_fe = wb['Flujo_Efectivo']
ws_fe['A1'] = "Estado de Flujo de Efectivo"
ws_fe['A1'].font = header_font
ws_fe['A1'].fill = header_fill
ws_fe.merge_cells('A1:N1')

for i, mes in enumerate(meses, start=2):
    ws_fe.cell(row=2, column=i, value=mes).font = header_font
    ws_fe.cell(row=2, column=i).fill = header_fill

# Flujo de Operaciones
ws_fe['A3'] = "FLUJO DE EFECTIVO DE OPERACIONES"
ws_fe['A3'].font = Font(bold=True, size=11, color="FFFFFF")
ws_fe['A3'].fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
ws_fe.merge_cells('A3:N3')

fe_operaciones = [
    ("Cobros a clientes", ["=Carga_Mensual!B16", "=Carga_Mensual!C16", "=Carga_Mensual!D16", "=Carga_Mensual!E16", "=Carga_Mensual!F16", "=Carga_Mensual!G16", "=Carga_Mensual!H16", "=Carga_Mensual!I16", "=Carga_Mensual!J16", "=Carga_Mensual!K16", "=Carga_Mensual!L16", "=Carga_Mensual!M16", "=SUM(B4:M4)"]),
    ("(-) Pago a proveedores", ["=Carga_Mensual!B17", "=Carga_Mensual!C17", "=Carga_Mensual!D17", "=Carga_Mensual!E17", "=Carga_Mensual!F17", "=Carga_Mensual!G17", "=Carga_Mensual!H17", "=Carga_Mensual!I17", "=Carga_Mensual!J17", "=Carga_Mensual!K17", "=Carga_Mensual!L17", "=Carga_Mensual!M17", "=SUM(B5:M5)"]),
    ("(-) Pago de remuneraciones", ["=Carga_Mensual!B18", "=Carga_Mensual!C18", "=Carga_Mensual!D18", "=Carga_Mensual!E18", "=Carga_Mensual!F18", "=Carga_Mensual!G18", "=Carga_Mensual!H18", "=Carga_Mensual!I18", "=Carga_Mensual!J18", "=Carga_Mensual!K18", "=Carga_Mensual!L18", "=Carga_Mensual!M18", "=SUM(B6:M6)"]),
    ("(-) Pago de gastos operacionales", ["=Carga_Mensual!B19", "=Carga_Mensual!C19", "=Carga_Mensual!D19", "=Carga_Mensual!E19", "=Carga_Mensual!F19", "=Carga_Mensual!G19", "=Carga_Mensual!H19", "=Carga_Mensual!I19", "=Carga_Mensual!J19", "=Carga_Mensual!K19", "=Carga_Mensual!L19", "=Carga_Mensual!M19", "=SUM(B7:M7)"]),
    ("(-) Pago de impuestos", ["=Carga_Mensual!B20", "=Carga_Mensual!C20", "=Carga_Mensual!D20", "=Carga_Mensual!E20", "=Carga_Mensual!F20", "=Carga_Mensual!G20", "=Carga_Mensual!H20", "=Carga_Mensual!I20", "=Carga_Mensual!J20", "=Carga_Mensual!K20", "=Carga_Mensual!L20", "=Carga_Mensual!M20", "=SUM(B8:M8)"]),
    ("FLUJO NETO DE OPERACIONES", ["=B4-B5-B6-B7-B8", "=C4-C5-C6-C7-C8", "=D4-D5-D6-D7-D8", "=E4-E5-E6-E7-E8", "=F4-F5-F6-F7-F8", "=G4-G5-G6-G7-G8", "=H4-H5-H6-H7-H8", "=I4-I5-I6-I7-I8", "=J4-J5-J6-J7-J8", "=K4-K5-K6-K7-K8", "=L4-L5-L6-L7-L8", "=M4-M5-M6-M7-M8", "=SUM(B9:M9)"])
]

for row, (concepto, formulas) in enumerate(fe_operaciones, start=4):
    ws_fe.cell(row=row, column=1, value=concepto).font = Font(bold=True if "FLUJO NETO" in concepto else False, size=10)
    for col, formula in enumerate(formulas, start=2):
        cell = ws_fe.cell(row=row, column=col, value=formula)
        cell.style = "currency"

# Flujo de Inversión
ws_fe['A10'] = "FLUJO DE EFECTIVO DE INVERSIÓN"
ws_fe['A10'].font = Font(bold=True, size=11, color="FFFFFF")
ws_fe['A10'].fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
ws_fe.merge_cells('A10:N10')

fe_inversion = [
    ("(-) Compra de activos fijos", ["=Carga_Mensual!B21", "=Carga_Mensual!C21", "=Carga_Mensual!D21", "=Carga_Mensual!E21", "=Carga_Mensual!F21", "=Carga_Mensual!G21", "=Carga_Mensual!H21", "=Carga_Mensual!I21", "=Carga_Mensual!J21", "=Carga_Mensual!K21", "=Carga_Mensual!L21", "=Carga_Mensual!M21", "=SUM(B11:M11)"]),
    ("(+) Venta de activos", ["=Carga_Mensual!B22", "=Carga_Mensual!C22", "=Carga_Mensual!D22", "=Carga_Mensual!E22", "=Carga_Mensual!F22", "=Carga_Mensual!G22", "=Carga_Mensual!H22", "=Carga_Mensual!I22", "=Carga_Mensual!J22", "=Carga_Mensual!K22", "=Carga_Mensual!L22", "=Carga_Mensual!M22", "=SUM(B12:M12)"]),
    ("FLUJO NETO DE INVERSIÓN", ["=B11+B12", "=C11+C12", "=D11+D12", "=E11+E12", "=F11+F12", "=G11+G12", "=H11+H12", "=I11+I12", "=J11+J12", "=K11+K12", "=L11+L12", "=M11+M12", "=SUM(B13:M13)"])
]

for row, (concepto, formulas) in enumerate(fe_inversion, start=11):
    ws_fe.cell(row=row, column=1, value=concepto).font = Font(bold=True if "FLUJO NETO" in concepto else False, size=10)
    for col, formula in enumerate(formulas, start=2):
        cell = ws_fe.cell(row=row, column=col, value=formula)
        cell.style = "currency"

# Flujo de Financiamiento
ws_fe['A14'] = "FLUJO DE EFECTIVO DE FINANCIAMIENTO"
ws_fe['A14'].font = Font(bold=True, size=11, color="FFFFFF")
ws_fe['A14'].fill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
ws_fe.merge_cells('A14:N14')

fe_financiamiento = [
    ("(+) Préstamos recibidos", ["=Carga_Mensual!B23", "=Carga_Mensual!C23", "=Carga_Mensual!D23", "=Carga_Mensual!E23", "=Carga_Mensual!F23", "=Carga_Mensual!G23", "=Carga_Mensual!H23", "=Carga_Mensual!I23", "=Carga_Mensual!J23", "=Carga_Mensual!K23", "=Carga_Mensual!L23", "=Carga_Mensual!M23", "=SUM(B15:M15)"]),
    ("(-) Pago de deuda", ["=Carga_Mensual!B24", "=Carga_Mensual!C24", "=Carga_Mensual!D24", "=Carga_Mensual!E24", "=Carga_Mensual!F24", "=Carga_Mensual!G24", "=Carga_Mensual!H24", "=Carga_Mensual!I24", "=Carga_Mensual!J24", "=Carga_Mensual!K24", "=Carga_Mensual!L24", "=Carga_Mensual!M24", "=SUM(B16:M16)"]),
    ("(+) Aportes de capital", ["=Carga_Mensual!B25", "=Carga_Mensual!C25", "=Carga_Mensual!D25", "=Carga_Mensual!E25", "=Carga_Mensual!F25", "=Carga_Mensual!G25", "=Carga_Mensual!H25", "=Carga_Mensual!I25", "=Carga_Mensual!J25", "=Carga_Mensual!K25", "=Carga_Mensual!L25", "=Carga_Mensual!M25", "=SUM(B17:M17)"]),
    ("(-) Retiros / dividendos", ["=Carga_Mensual!B26", "=Carga_Mensual!C26", "=Carga_Mensual!D26", "=Carga_Mensual!E26", "=Carga_Mensual!F26", "=Carga_Mensual!G26", "=Carga_Mensual!H26", "=Carga_Mensual!I26", "=Carga_Mensual!J26", "=Carga_Mensual!K26", "=Carga_Mensual!L26", "=Carga_Mensual!M26", "=SUM(B18:M18)"]),
    ("FLUJO NETO DE FINANCIAMIENTO", ["=B15-B16+B17-B18", "=C15-C16+C17-C18", "=D15-D16+D17-D18", "=E15-E16+E17-E18", "=F15-F16+F17-F18", "=G15-G16+G17-G18", "=H15-H16+H17-H18", "=I15-I16+I17-I18", "=J15-J16+J17-J18", "=K15-K16+K17-K18", "=L15-L16+L17-L18", "=M15-M16+M17-M18", "=SUM(B19:M19)"])
]

for row, (concepto, formulas) in enumerate(fe_financiamiento, start=15):
    ws_fe.cell(row=row, column=1, value=concepto).font = Font(bold=True if "FLUJO NETO" in concepto else False, size=10)
    for col, formula in enumerate(formulas, start=2):
        cell = ws_fe.cell(row=row, column=col, value=formula)
        cell.style = "currency"

# Variación Neta del Efectivo
ws_fe['A20'] = "VARIACIÓN NETA DEL EFECTIVO"
ws_fe['A20'].font = Font(bold=True, size=12, color="FFFFFF")
ws_fe['A20'].fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
ws_fe.merge_cells('A20:N20')

for col in range(2, 15):
    ws_fe.cell(row=20, column=col).value = f"=B9+B13+B19"
    ws_fe.cell(row=20, column=col).style = "currency"

# Efectivo Inicial y Final
ws_fe['A21'] = "Efectivo inicial del período"
ws_fe['A21'].font = body_font
for col in range(2, 15):
    if col == 2:
        ws_fe.cell(row=21, column=col, value="=Carga_Mensual!B27").style = "currency"
    else:
        ws_fe.cell(row=21, column=col, value=f"=B21+B20").style = "currency"

ws_fe['A22'] = "Efectivo final del período"
ws_fe['A22'].font = body_font
for col in range(2, 15):
    ws_fe.cell(row=22, column=col, value=f"=B21").style = "currency"

# Aplicar bordes
apply_border(ws_fe, 'A1:N22')

# Congelar encabezados
freeze_panes(ws_fe, 3, 2)

# Ajustar anchos
ws_fe.column_dimensions['A'].width = 30
for col in range(2, 15):
    ws_fe.column_dimensions[get_column_letter(col)].width = 15

# Hoja 8: Cambios_Patrimonio

# Hoja 9: Ratios_KPI
# Con fórmulas de ratios

# Hoja 10: Dashboard
ws_dash = wb['Dashboard']
ws_dash['A1'] = "Dashboard Ejecutivo - Estados Financieros Completos Premium"
ws_dash['A1'].font = Font(bold=True, size=18, color="1F497D")
ws_dash['A1'].alignment = Alignment(horizontal="center")
ws_dash.merge_cells('A1:I1')

# Título de empresa
ws_dash['A2'] = f"=Parametros!B5"
ws_dash['A2'].font = Font(bold=True, size=14, color="4F81BD")
ws_dash['A2'].alignment = Alignment(horizontal="center")
ws_dash.merge_cells('A2:I2')

# Sección 1: Indicadores Clave
ws_dash['A4'] = "Indicadores Clave de Rendimiento (KPI)"
ws_dash['A4'].font = header_font
ws_dash['A4'].fill = header_fill
ws_dash.merge_cells('A4:I4')

# KPI 1: Ventas Totales Anuales
ws_dash['A6'] = "Ventas Totales Anuales"
ws_dash['B6'] = "=Estado_Resultados!N5"  # Ventas netas totales
ws_dash['B6'].style = "currency"
ws_dash['B6'].font = Font(bold=True, size=12, color="006600")

# KPI 2: Utilidad Neta Anual
ws_dash['A7'] = "Utilidad Neta Anual"
ws_dash['B7'] = "=Estado_Resultados!N17"  # Utilidad neta total
ws_dash['B7'].style = "currency"
ws_dash['B7'].font = Font(bold=True, size=12)

# KPI 3: Margen de Utilidad
ws_dash['A8'] = "Margen de Utilidad (%)"
ws_dash['B8'] = "=IF(B6>0, B7/B6, 0)"
ws_dash['B8'].style = "percentage"
ws_dash['B8'].font = Font(bold=True, size=12)

# KPI 4: Liquidez (Ratio Corriente)
ws_dash['A9'] = "Ratio Corriente"
ws_dash['B9'] = "=Balance_General!N8/Balance_General!N22"  # Activos corrientes / Pasivos corrientes
ws_dash['B9'].number_format = "0.00"
ws_dash['B9'].font = Font(bold=True, size=12)

# KPI 5: Endeudamiento
ws_dash['A10'] = "Ratio de Endeudamiento (%)"
ws_dash['B10'] = "=Balance_General!N27/Balance_General!N15"  # Total pasivos / Total activos
ws_dash['B10'].style = "percentage"
ws_dash['B10'].font = Font(bold=True, size=12)

# Aplicar colores condicionales a KPIs
from openpyxl.formatting.rule import Rule
red_fill = PatternFill(bgColor="FFC7CE")
green_fill = PatternFill(bgColor="C6EFCE")

# Margen de utilidad: verde si >10%, rojo si <5%
ws_dash.conditional_formatting.add('B8:B8', Rule(type='cellIs', operator='greaterThan', formula=['0.1'], fill=green_fill))
ws_dash.conditional_formatting.add('B8:B8', Rule(type='cellIs', operator='lessThan', formula=['0.05'], fill=red_fill))

# Ratio corriente: verde si >1.5, rojo si <1
ws_dash.conditional_formatting.add('B9:B9', Rule(type='cellIs', operator='greaterThan', formula=['1.5'], fill=green_fill))
ws_dash.conditional_formatting.add('B9:B9', Rule(type='cellIs', operator='lessThan', formula=['1'], fill=red_fill))

# Endeudamiento: verde si <50%, rojo si >70%
ws_dash.conditional_formatting.add('B10:B10', Rule(type='cellIs', operator='lessThan', formula=['0.5'], fill=green_fill))
ws_dash.conditional_formatting.add('B10:B10', Rule(type='cellIs', operator='greaterThan', formula=['0.7'], fill=red_fill))

# Sección 2: Gráficos de Tendencia Mensual
ws_dash['D6'] = "Tendencia de Ventas Mensuales"
ws_dash['D6'].font = header_font
ws_dash['D6'].fill = header_fill
ws_dash.merge_cells('D6:I6')

# Datos para gráfico de ventas (referencias a Estado_Resultados)
for i, mes in enumerate(meses[:-1], start=7):  # Excluir Total Anual
    ws_dash.cell(row=i, column=4, value=mes[:3])  # Mes abreviado
    ws_dash.cell(row=i, column=5, value=f"=Estado_Resultados!{get_column_letter(i-4)}5")  # Ventas netas

# Crear gráfico de barras para ventas
from openpyxl.chart import BarChart, Reference, Series
chart_ventas = BarChart()
chart_ventas.type = "col"
chart_ventas.style = 10
chart_ventas.title = "Ventas Mensuales"
chart_ventas.y_axis.title = 'Monto (CLP)'
chart_ventas.x_axis.title = 'Mes'

data = Reference(ws_dash, min_col=5, min_row=7, max_row=18)
cats = Reference(ws_dash, min_col=4, min_row=7, max_row=18)
chart_ventas.add_data(data, titles_from_data=True)
chart_ventas.set_categories(cats)
ws_dash.add_chart(chart_ventas, "D8")

# Gráfico de Utilidad Neta
ws_dash['D20'] = "Tendencia de Utilidad Neta Mensual"
ws_dash['D20'].font = header_font
ws_dash['D20'].fill = header_fill
ws_dash.merge_cells('D20:I20')

for i, mes in enumerate(meses[:-1], start=21):
    ws_dash.cell(row=i, column=4, value=mes[:3])
    ws_dash.cell(row=i, column=5, value=f"=Estado_Resultados!{get_column_letter(i-19)}17")

chart_utilidad = BarChart()
chart_utilidad.type = "col"
chart_utilidad.style = 11
chart_utilidad.title = "Utilidad Neta Mensual"
chart_utilidad.y_axis.title = 'Monto (CLP)'
chart_utilidad.x_axis.title = 'Mes'

data_util = Reference(ws_dash, min_col=5, min_row=21, max_row=32)
cats_util = Reference(ws_dash, min_col=4, min_row=21, max_row=32)
chart_utilidad.add_data(data_util, titles_from_data=True)
chart_utilidad.set_categories(cats_util)
ws_dash.add_chart(chart_utilidad, "D22")

# Sección 3: Composición del Balance General
ws_dash['A12'] = "Composición del Balance General"
ws_dash['A12'].font = header_font
ws_dash['A12'].fill = header_fill
ws_dash.merge_cells('A12:C12')

# Datos para gráfico circular
ws_dash['A14'] = "Activos Corrientes"
ws_dash['B14'] = "=Balance_General!N8"  # Total activos corrientes
ws_dash['A15'] = "Activos No Corrientes"
ws_dash['B15'] = "=Balance_General!N14"  # Total activos no corrientes
ws_dash['A16'] = "Pasivos Corrientes"
ws_dash['B16'] = "=Balance_General!N22"  # Total pasivos corrientes
ws_dash['A17'] = "Pasivos No Corrientes"
ws_dash['B17'] = "=Balance_General!N26"  # Total pasivos no corrientes
ws_dash['A18'] = "Patrimonio"
ws_dash['B18'] = "=Balance_General!N35"  # Total patrimonio

# Gráfico circular para composición del balance
from openpyxl.chart import PieChart
pie_chart = PieChart()
pie_chart.title = "Composición del Balance General"
pie_chart.style = 13

pie_data = Reference(ws_dash, min_col=2, min_row=14, max_row=18)
pie_labels = Reference(ws_dash, min_col=1, min_row=14, max_row=18)
pie_chart.add_data(pie_data, titles_from_data=True)
pie_chart.set_categories(pie_labels)
ws_dash.add_chart(pie_chart, "A20")

# Sección 4: Flujo de Efectivo
ws_dash['A25'] = "Resumen de Flujo de Efectivo Anual"
ws_dash['A25'].font = header_font
ws_dash['A25'].fill = header_fill
ws_dash.merge_cells('A25:C25')

ws_dash['A27'] = "Flujo de Operaciones"
ws_dash['B27'] = "=Flujo_Efectivo!N9"
ws_dash['B27'].style = "currency"

ws_dash['A28'] = "Flujo de Inversión"
ws_dash['B28'] = "=Flujo_Efectivo!N13"
ws_dash['B28'].style = "currency"

ws_dash['A29'] = "Flujo de Financiamiento"
ws_dash['B29'] = "=Flujo_Efectivo!N19"
ws_dash['B29'].style = "currency"

ws_dash['A30'] = "Variación de Caja"
ws_dash['B30'] = "=Flujo_Efectivo!N20"
ws_dash['B30'].style = "currency"

# Gráfico de cascada para flujo de efectivo
from openpyxl.chart import BarChart
waterfall_chart = BarChart()
waterfall_chart.type = "col"
waterfall_chart.style = 12
waterfall_chart.title = "Flujo de Efectivo Anual"
waterfall_chart.y_axis.title = 'Monto (CLP)'

wf_data = Reference(ws_dash, min_col=2, min_row=27, max_row=30)
wf_cats = Reference(ws_dash, min_col=1, min_row=27, max_row=30)
waterfall_chart.add_data(wf_data, titles_from_data=True)
waterfall_chart.set_categories(wf_cats)
ws_dash.add_chart(waterfall_chart, "A32")

# Ajustar anchos de columna
ws_dash.column_dimensions['A'].width = 25
ws_dash.column_dimensions['B'].width = 20
ws_dash.column_dimensions['C'].width = 15
ws_dash.column_dimensions['D'].width = 10
ws_dash.column_dimensions['E'].width = 15

# Aplicar bordes a secciones
apply_border(ws_dash, 'A4:B10')
apply_border(ws_dash, 'A12:B18')
apply_border(ws_dash, 'A25:B30')

# Finalmente, guardar
wb.save("Sistema_Financiero_PyME.xlsx")
print("Archivo Excel 'Sistema_Financiero_PyME.xlsx' creado exitosamente.")